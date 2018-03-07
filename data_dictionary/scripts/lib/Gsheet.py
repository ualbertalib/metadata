from __future__ import print_function
import httplib2
import os
import pprint
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
#import apiclient
#import discovery
from googleapiclient import discovery
from oauth2client import client
from oauth2client import tools
from oauth2client.file import Storage

try:
    import argparse
    flags = argparse.ArgumentParser(parents=[tools.argparser]).parse_args()
except ImportError:
    flags = None

# If modifying these scopes, delete your previously saved credentials
# at ~/.credentials/sheets.googleapis.com-python-quickstart.json
SCOPES = 'https://www.googleapis.com/auth/spreadsheets'
CLIENT_SECRET_FILE = 'client_secret.json'
APPLICATION_NAME = 'Google Sheets API Quickstart'


def get_credentials():

    home_dir = os.path.expanduser('~')
    credential_dir = os.path.join(home_dir, '.credentials')
    if not os.path.exists(credential_dir):
        os.makedirs(credential_dir)
    credential_path = os.path.join(credential_dir,
                                   'sheets.googleapis.com-python-quickstart.json')

    store = Storage(credential_path)
    credentials = store.get()
    if not credentials or credentials.invalid:
        flow = client.flow_from_clientsecrets(CLIENT_SECRET_FILE, SCOPES)
        flow.user_agent = APPLICATION_NAME
        if flags:
            credentials = tools.run_flow(flow, store, flags)
        else: # Needed only for compatibility with Python 2.6
            credentials = tools.run(flow, store)
        print('Storing credentials to ' + credential_path)
    return credentials

def google_generate():
    
    credentials = get_credentials()
    http = credentials.authorize(httplib2.Http())
    discoveryUrl = ('https://sheets.googleapis.com/$discovery/rest?'
                    'version=v4')
    service = discovery.build('sheets', 'v4', http=http,
                              discoveryServiceUrl=discoveryUrl)

    spreadsheetId = '1ThPCrBePu74BchbgsXiZpXgEQzVsGsMVyc_gn9VT0ns'

    sheets = service.spreadsheets().get(spreadsheetId=spreadsheetId)
    sheet = sheets.execute()
    length = len(sheet['sheets'])

    addtemp = {
    "requests": [
    {
      "addSheet": {
        "properties": {
          "title": "TempSheet",
          "sheetId" : "123456"
        }
      }
    }
  ]

    }
    request = service.spreadsheets().batchUpdate(spreadsheetId=spreadsheetId, body=addtemp)
    response = request.execute()

    for index in range(0 , length):
        sheetId = sheet['sheets'][index]['properties']['sheetId']

        deletesheets = {
            "requests": [
                {
                  "deleteSheet": {
                    "sheetId": sheetId
                    }
                }
            ]
        }
        request = service.spreadsheets().batchUpdate(spreadsheetId=spreadsheetId, body=deletesheets)
        response = request.execute()

    wb = Workbook()
    #../../profiles.xlsx
    wb = openpyxl.load_workbook('data_dictionary/profiles.xlsx')
    names = wb.get_sheet_names()
    for sid, sheetName in enumerate(names):
        add = {
            "requests": [
                {
                "addSheet": {
                    "properties": {
                        "title": sheetName,
                        "sheetId" : int(sid + 1)
                    }
                }
            }
            ]
        }
        request = service.spreadsheets().batchUpdate(spreadsheetId=spreadsheetId, body=add)
        response = request.execute()
        rangeName = str(sheetName+'!A1:AA100')
        ws = wb[sheetName]
        val = []
        final = []
        extras = []
        for i in range(1, ws.max_row+1):
            final.append(val)
            val = []
            for j in range(1, ws.max_column+1):
                value = ws.cell(row=i, column=j).value
                if value == 'acceptedValues':
                    extras.append(j)
                val.append(value)       
        del final[0]
        values = final
        #print (values)
        body = {
          'values': values
        }
        result = service.spreadsheets().values().update(spreadsheetId=spreadsheetId, range=rangeName, valueInputOption='USER_ENTERED', body=body).execute()

        resize = {
            "requests": [
                {
                    "autoResizeDimensions": {
                        "dimensions": {
                            "sheetId": int(sid + 1),
                            "dimension": "COLUMNS",
                            "startIndex": 0,
                            "endIndex": 25
                        }
                    }
                }
            ]
        }
        request = service.spreadsheets().batchUpdate(spreadsheetId=spreadsheetId, body=resize)
        response = request.execute()

        for z, number in enumerate(extras):
            number = number - z
            deleteCols = {
                "requests": [
                    {
                        "deleteDimension": {
                            "range": {
                                "sheetId": int(sid + 1),
                                "dimension": "COLUMNS",
                                "startIndex": int(number - 1),"endIndex": int(number)
                                }
                        }
                    },
                ],
            }
            request = service.spreadsheets().batchUpdate(spreadsheetId=spreadsheetId, body=deleteCols)
            response = request.execute()

        protecte = {
            "requests": [
                {
                    "addProtectedRange": {
                        "protectedRange": {
                            "range": {
                                "sheetId": int(sid + 1),
                                "startRowIndex": 0,
                                "endRowIndex": 99,
                                "startColumnIndex": 0,
                                "endColumnIndex": 25,
                            },
                            "description": "Protecting DD",
                            "warningOnly": False,
                            "editors": {
                                "users": [
                                    "danoosh@ualberta.ca",
                                    "zschoenb@ualberta.ca",
                                    "sharon.farnel@ualberta.ca"
                                ]
                            }
                        }
                    }
                }
            ]
        }
        request = service.spreadsheets().batchUpdate(spreadsheetId=spreadsheetId, body=protecte)
        response = request.execute()

    deleteTempSheet = {
        "requests": [
            {
                "deleteSheet": {
                    "sheetId": "123456"
                }
            }
        ]
    }
    request = service.spreadsheets().batchUpdate(spreadsheetId=spreadsheetId, body=deleteTempSheet)
    response = request.execute()
    
if __name__ == '__main__':
    google_generate()
