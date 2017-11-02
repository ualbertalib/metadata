import os
import json
import openpyxl
from openpyxl import Workbook
from config import namespaces, ignore
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

class Excel:

    def __init__(self, profile):
        self.profile = profile
        self.headings = []
        self.generate()
    
    def generate(self):      
        with open('../../profiles/' + self.profile + '/' + 'profile.json') as data_file:
            wb = Workbook()
            wb = openpyxl.load_workbook('../../profiles.xlsx')
            self.writeNamespaces(wb)
            sheet1 = wb.create_sheet(title = 'Accepted_Values')
            sheet = wb.create_sheet(title = self.profile)
            data = json.load(data_file)
            index = 1
            startrow = 2
            for i, key in enumerate(data.keys()):
                for ig in namespaces:
                    if ig['uri'] in key:
                        sheet['A' + str(i+2)] = key.replace(ig['uri'],ig['prefix'] + ' : ')
                        break
                    else:
                        sheet['A' + str(i+2)] = key
                self.headings = self.getHeadings(data, key)
                for x, title in enumerate(self.headings):
                    col = get_column_letter(x+2)
                    sheet[str(col)+str(1)] = title.replace('http://terms.library.ualberta.ca/', '')                    
                    if title in data[key].keys():
                        for items in data[key][title]:
                            sheet[str(col)+str(i+2)] = str(items)
                        for namespace in namespaces:
                            # remove namespace uris
                            if str(namespace['uri']) in str(data[key][title]):
                                for items in data[key][title]:
                                    cell = str(items).replace(namespace['uri'], namespace['prefix'] + ' : ')
                                    sheet[str(col)+str(i+2)] = cell 
                        try:
                            if type(data[key][title][0]) is dict:
                                # this will write the list of accepted values in a multiline cell
                                '''style = Alignment(wrap_text=True)
                                celll = sheet[str(col)+str(i+2)]
                                celll.alignment = styledata[key][title]
                                length = len(data[key][title])
                                cellvalue = []
                                for subs in range(0, length-1):
                                    for ke in data[key][title][subs].keys():
                                        cellvalue.append(str(ke) + " : " + str(data[key][title][subs][ke]) + " | ")
                                    cellvalue.append("\n")
                                    string= ""
                                    for item in cellvalue:
                                        string = string + str(item) + " "
                                    sheet[str(col)+str(i+2)] = str(string)'''
                                # this will write the accepted values in a new sheet
                                ws = wb.worksheets[1]
                                ws['A'+str(startrow)] = str(key)
                                for h, ke in enumerate(data[key][title][0].keys()):
                                    col = get_column_letter(h+2)
                                    ws[str(col)+str(1)] = str(ke)
                                    length = len(data[key][title])
                                    for subs in range(0, length):
                                        ws[str(col)+str(startrow + subs)] = str(data[key][title][subs][ke])
                                startrow = startrow + length
                        except Exception:
                            continue       
                    else: 
                        sheet[str(col)+str(i+2)] = "N/A"
            names = wb.get_sheet_names()
            for name in names:
                if name in ['community', 'collection', 'generic', 'thesis', 'Namespaces', 'Accepted_Values']:
                    continue
                else:
                    std=wb.get_sheet_by_name(name)
                    wb.remove_sheet(std)
            sheet.column_dimensions["A"].width = 60.0
            for colname in range(2, 19):
                column = get_column_letter(colname)
                sheet.column_dimensions[column].width = 30.0
            sheet1.column_dimensions["A"].width = 60.0
            wb.save('../../profiles.xlsx')
            
    def getHeadings(self, data, key):
        self.key = key
        self.data = data
        for ki in data[self.key].keys():
            if ki not in self.headings:
                self.headings.append(ki)
        return self.headings
    
    def writeNamespaces(self, wb):
        self.wb = wb
        self.sheet = self.wb.create_sheet(title = 'Namespaces')
        for z, names in enumerate(namespaces):
            self.sheet['A'+str(z+1)] = str(names['prefix'])
            self.sheet['B'+str(z+1)] = str(names['uri'])
