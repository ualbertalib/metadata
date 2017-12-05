import os
import openpyxl
from excel import Excel
from openpyxl import Workbook


def excelGen():
    os.remove('data_dictionary/profiles.xlsx')
    #../../profiles.xlsx
    wb = Workbook()
    ws = wb.active
    ws['A1'] = ''
    wb.save('data_dictionary/profiles.xlsx')
    #../../profiles.xlsx
    for profile in ['community', 'collection', 'generic', 'thesis', 'oai_pmh', 'oai_etdms']:
    	Excel(profile)
        
if __name__ == "__main__":
    excelGen()


class Excelx:

    def __init__(self, profile):
        self.profile = profile
        self.headings= []
        
    def generate(self):
        os.remove(self.profile + '.xlsx')
        wb = Workbook()
        ws = wb.active
        ws['A1'] = ''
        wb.save(self.profile + '.xlsx')
        with open(self.profile + '.json') as data_file:
            wb = Workbook()
            wb = openpyxl.load_workbook(self.profile + '.xlsx')
            sheet = wb.active
            data = json.load(data_file)
            for i, key in enumerate(data.keys()):
                sheet['A' + str(i+2)] = key
                self.headings = self.getHeadings(key)
                for j, k in enumerate(data[key].keys()):
                    col = get_column_letter(j+2)
                    sheet[str(col)+str(1)] = k.replace('http://terms.library.ualberta.ca/', '')
                    sheet[str(col)+str(i+2)] = str(data[key][k])
            wb.save(self.profile + '.xlsx')      

