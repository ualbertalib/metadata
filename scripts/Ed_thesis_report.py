import os
import json
import openpyxl
from fuzzywuzzy import fuzz
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

wb = Workbook()
wb = openpyxl.load_workbook('educationtheses.xlsx')
sheet = wb.active
wb1 = openpyxl.load_workbook('Thesis.xlsx')
sheet1 = wb1.active
val = []
final = []
for i in range(1, sheet.max_row+1):
    final.append(val)
    val = []
    for j in range(1, sheet.max_column+1):
        value = sheet.cell(row=i, column=j).value
        val.append(value)       
del final[0]
val1 = []
final1 = []
for i in range(1, sheet1.max_row+1):
    final1.append(val1)
    val1 = []
    for j in range(1, sheet1.max_column+1):
        value1 = sheet1.cell(row=i, column=j).value
        val1.append(value1)       
del final1[0] 
with open('test.tsv', 'a') as file:
	z = 0
	file.write("name in catalogue" + "\t" + "title in catalogue" + "\t" + "name in ERA" + "\t" + "title in ERA" + "\t" + "department in catalogue" + "\t" + "department in ERA" + "\t" + "year in catalogue" + "\t" + "year in ERA" + "\t" + "degeree in catalogue" + "\t" + "degree in ERA" + "\t" + "microfilm?" + "\n")
	for i in final:
		for j in final1:
			f = 0
			auth = fuzz.token_sort_ratio(j[2], i[0])
			title = fuzz.token_sort_ratio(j[4], i[1])
			if auth > 95 and title > 95:
				print (j[2] + " " + i[0] + " " + str(auth))
				z = z + 1
				f = 1
				file.write(i[0].encode('utf-8') + "\t" + i[1].encode('utf-8') + "\t" + j[2].encode('utf-8') + "\t" + j[4].encode('utf-8') + "\t" + i[5].encode('utf-8') + "\t" + j[1].encode('utf-8') + "\t" + str(i[3]) + "\t" + str(j[5]) + "\t" + i[4].encode('utf-8') + "\t" + j[3].encode('utf-8') + "\t" + str(i[2]) + "\n")


				#sheet3['A' + str(z)] = i[0]
				#sheet3['B' + str(z)] = i[1]
				#sheet3['C' + str(z)] = j[2]
				#sheet3['D' + str(z)] = j[4]
				#sheet3['E' + str(z)] = i[2]
				#sheet3['F' + str(z)] = i[3]
				print (z)
				print (i[0] + " title: " + i[1] + "  ---   " + j[2] + "  title: " + j[4])
				break
		if f == 0:
			print(i[0] + " title: " + i[1] + "  ---   " + j[2] + "  title: " + j[4])
			z = z + 1
			file.write(i[0].encode('utf-8') + "\t" + i[1].encode('utf-8') + "\t" + "Not in ERA" + "\t" + "Not in ERA" + "\t" + i[5].encode('utf-8') + "\t" + "Not in ERA" + "\t" + str(i[3]) + "\t" + "Not in ERA" + "\t" + i[4].encode('utf-8') + "\t" + "Not in ERA" + "\t" + str(i[2]) +"\n")
		

		#sheet3['A' + str(z)] = i[0]
		#sheet3['B' + str(z)] = i[1]
		#sheet3['C' + str(z)] = "Not in ERA"
		#sheet3['D' + str(z)] = "Not in ERA"
		#sheet3['E' + str(z)] = i[2]
		#sheet3['F' + str(z)] = i[3]

