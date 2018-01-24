import openpyxl
from openpyxl import Workbook
from pymarc import MARCReader
from pymarc import Record, Field
from pymarc import XmlHandler
from pymarc import record_to_xml
from pymarc import XMLWriter
import glob
import os
import time
from datetime import datetime

ts = datetime.fromtimestamp(time.time()).strftime('%H:%M:%S')

path = 'extracts/'
for filename in glob.glob(os.path.join(path, '*.xlsx')):
	folder = filename.replace('.xlsx', '').replace(path, '')
	wb = Workbook()
	wb = openpyxl.load_workbook(filename)
	names = wb.get_sheet_names()
	li = {}
	ws = wb[names[0]]
	for i in range(1, ws.max_row+1):
		if str(ws.cell(row=i, column=2).value) == 'None' and str(ws.cell(row=i, column=3).value) == 'None':
			continue
		else:
			key = str(ws.cell(row=i, column=1).value.replace('http://example.org/', '').split('#')[0])
			lc = str(ws.cell(row=i, column=2).value)
			# append this url in openRefine process
			vf = 'http://viaf.org/viaf/' + str(ws.cell(row=i, column=3).value)
			li[key] = {'lc' : lc, 'vf' : vf}
	#print (li)
	with open('marc/' + folder + '.mrc', 'rb') as fh:
		reader = MARCReader(fh, to_unicode=True, force_utf8=False, utf8_handling='ignore')
		ind = 0
		for record in reader:
			ind += 1
			print (ind)
			for f in record.get_fields('001'):
				f = str(f).replace('=001  ', '')
	   			for i in li.keys():
	   				if i == str(f):
	   					if li[i]['lc'] == 'None':
		   					record.add_field(
							    Field(
							        tag = '758',
							        indicators = ['0','1'],
							        subfields = [
							            'b', li[i]['vf']
							        ]))
		   				elif li[i]['vf'] == 'http://viaf.org/viaf/None':
		   					record.add_field(
							    Field(
							        tag = '758',
							        indicators = ['0','1'],
							        subfields = [
							            'a', li[i]['lc']
							        ]))
		   				else:
		   					record.add_field(
							    Field(
							        tag = '758',
							        indicators = ['0','1'],
							        subfields = [
							            'a', li[i]['lc'],
							            'b', li[i]['vf'],
							        ]))
		   				if not os.path.exists(folder):
		   					os.makedirs(folder)
		   					os.makedirs(folder + '/mrc')
						out = open(folder + '/mrc/' + f + '.mrc', 'wb')
						out.write(record.as_marc())
						out.close()
						out = open(folder + '/' + 'mrc.mrc', 'a+')
						out.write(record.as_marc())
						out.close()

tf = datetime.fromtimestamp(time.time()).strftime('%H:%M:%S')
print("walltime:", datetime.strptime(tf, '%H:%M:%S') - datetime.strptime(ts, '%H:%M:%S'))
